from openpyxl import load_workbook
import pandas as pd
writer2 = pd.ExcelWriter('Tableau stock modele_resultat.xlsx')
workbook = load_workbook(filename="Tableau stock modele.xlsx")
sheets=workbook.sheetnames;
prod,vente,transfert=0,0,0
for i in sheets:
    sheetR = workbook[i]
    valuesCode=[]
    valuesDes=[]
    valuesRes=[]
    for value in sheetR.iter_rows(min_row=2,values_only=True):
        valuesCode.append(value[0])
        valuesDes.append(value[1])
        valuesRes.append(value[2]+prod-vente-transfert)
    sheet = pd.DataFrame({'Code du produit':valuesCode,'Désignation du produit':valuesDes,'Stock':valuesRes})
    sheet.to_excel(writer2, sheet_name = i, index = False)
    prod=value[2]
    vente=value[3]
    transfert=value[4]
writer2.save()  
writer2.close()  
